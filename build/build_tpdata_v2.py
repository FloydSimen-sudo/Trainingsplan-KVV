#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Trainingsplan KVV — Excel -> tp-data.js  (Version 2, Neuaufbau)
Liest alle echten Quelldateien aus dem Testspace-Ordner und baut eine
tp-data.js im Schema, das die App (index.html) erwartet.

Kein Wert wird erfunden: fehlt eine Angabe in Excel, wird null geschrieben
und die App zeigt "-" an.
"""
import openpyxl, json, re, subprocess, sys, glob, os
from datetime import date

BASE = '/mnt/user-data/uploads/Testspace/'
OUT = '/home/claude/tpapp/design/tp-data.js'

def resolve_file(*candidates):
    """Nimmt den ersten existierenden Pfad; wenn ein Kandidat nicht exakt
    existiert, wird VOR dem nächsten Kandidaten erst per Glob nach einer
    OneDrive-Konfliktkopie im selben Ordner gesucht (z.B. 'Name 2.xlsx').
    Wichtig: das muss pro Kandidat passieren (nicht erst ganz am Ende),
    sonst gewinnt ein veralteter Fallback-Pfad (z.B. eine liegen gebliebene
    Kopie im alten Wurzelordner) fälschlich gegen eine Konfliktkopie im
    eigentlich richtigen, neueren Ordner (Bug bei Jakob Burtscher 28./30.08.:
    root-Datei 'Jakob Burtscher 26_27.xlsx' war ein alter Stand ohne die
    neuen Trainingsinhalte, wurde aber vor der echten Konfliktkopie in
    Einzelpläne/ gefunden)."""
    for cand in candidates:
        if os.path.exists(cand):
            return cand
        base_dir = os.path.dirname(cand)
        stem = os.path.splitext(os.path.basename(cand))[0]
        prefix = re.sub(r'\s*\d*$', '', stem).strip()
        hits = glob.glob(os.path.join(base_dir, prefix + '*.xlsx')) if os.path.isdir(base_dir) else []
        if hits:
            return hits[0]
    return candidates[0]

# ---------------------------------------------------------------- Kader ----

KADER_FILE = BASE + 'Zusatzinfos/Kaderaufstellung – Kopie.xlsx'

def norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower()) if s else ''

def load_kader():
    wb = openpyxl.load_workbook(KADER_FILE, data_only=True)
    ws = wb['Tabelle1']
    groups = []          # [{id, coaches, athletes:[{n,wk}]}]
    cur = None
    for r in range(4, ws.max_row + 1):
        vor = ws.cell(row=r, column=2).value
        nach = ws.cell(row=r, column=3).value
        wk27 = ws.cell(row=r, column=5).value
        grp = ws.cell(row=r, column=6).value
        trainer = ws.cell(row=r, column=7).value
        if not vor or str(vor).strip() in ('', '\xa0'):
            continue
        name = f"{str(vor).strip()} {str(nach).strip()}"
        if grp and str(grp).strip():
            gid = str(grp).strip()
            cur = {'id': gid, 'coaches': (str(trainer).strip() if trainer else ''), 'athletes': []}
            groups.append(cur)
        if cur is not None:
            cur['athletes'].append({'n': name, 'wk': str(wk27).strip() if wk27 else None})
    return groups

# --------------------------------------------------------- Übungssammlung --

def load_ex():
    wb = openpyxl.load_workbook(BASE + 'Zusatzinfos/Übungssammlung.xlsx', data_only=True)
    ws = wb['Tabelle1']
    ex = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=3).value
        if not name:
            continue
        ex[str(name).strip()] = {
            'min': ws.cell(row=r, column=4).value,
            'sets': ws.cell(row=r, column=5).value,
            'vol': ws.cell(row=r, column=6).value,
            'pause': ws.cell(row=r, column=7).value,
            'how': ws.cell(row=r, column=8).value,
        }
    return ex

CATALOG = None

def build_catalog(ex):
    cat = []
    for name, v in ex.items():
        if isinstance(v.get('min'), (int, float)):
            cat.append((name, v['min']))
    return cat

def words(s):
    return [w for w in re.findall(r'[a-zäöüß]+', str(s).lower()) if len(w) >= 3]

def fuzzy_match(text, catalog, ex_full):
    """Findet den ähnlichsten Katalogeintrag (>=50% Wortüberdeckung) und
    liefert dessen komplette Übungssammlung-Zeile (min/sets/vol/pause)."""
    tw = words(text)
    if not tw:
        return None
    best, best_score = None, 0
    for name, _dauer in catalog:
        cw = words(name)
        if not cw:
            continue
        hits = sum(1 for c in cw if any(c in t or t in c for t in tw))
        score = hits / len(cw)
        if score > best_score:
            best_score, best = score, name
    if best_score >= 0.5 and best in ex_full:
        rec = ex_full[best]
        return {'name': best, 'min': rec.get('min'), 'sets': rec.get('sets'), 'vol': rec.get('vol'), 'pause': rec.get('pause')}
    return None

# --------------------------------------------------------------- Content --
# Feste Kategorien laut Floyd (immer alle 8, unabhängig von Einzel-/Gruppenplan):
CAT_ORDER = ['Theorie', 'Aufwärmen', 'Motorik', 'Technikfokus', 'Technik/Taktik', 'Physis Klettern', 'Athletik + Mentales', 'Spezial']

# Individuelle Einheitenplanung: Zeilen-Offsets relativ zum Session-Basiswert
# (base=4 für Session 1, base=19 für Session 2) — siehe Kopfzeilen-Dump.
IND_CAT_ROWS = {
    'Theorie': [3], 'Aufwärmen': [4], 'Motorik': [5], 'Technikfokus': [6],
    'Technik/Taktik': [7, 8], 'Physis Klettern': [9, 10], 'Athletik + Mentales': [11],
    'Spezial': [13],
}
IND_NOTIZ_OFFSETS = [12, 14]

# 30.08.: Übergangs-Fallback fürs alte Einzel-Sheet "Einheitenplanung" (ohne
# "26_27"). Floyd befüllt manche Wochen (z.B. KW36) noch dort statt im neuen
# Sheet — für Tage, die im neuen Sheet komplett leer sind, wird der Inhalt
# automatisch aus dem alten Sheet nachgezogen (1 Session statt 2, gleiche
# 8 Kategorien, Skill->Technik/Taktik, Physis Wand->Physis Klettern,
# Physis Boden->Athletik + Mentales, Besprechung->Theorie — Mapping wie beim
# Statistik-Saison-Parser für Gruppen bereits etabliert).
OLD_IND_CAT_ROWS = {
    'Theorie': [6], 'Aufwärmen': [7, 8], 'Motorik': [], 'Technikfokus': [9],
    'Technik/Taktik': [10, 11, 12], 'Physis Klettern': [13, 14, 15],
    'Athletik + Mentales': [16, 17, 18, 19, 20], 'Spezial': [21, 22, 23],
}
OLD_IND_NOTIZ_ROWS = [27]
OLD_IND_ORT_ROW = 5

def build_old_individual_index(wb):
    if 'Einheitenplanung' not in wb.sheetnames:
        return None
    ws = wb['Einheitenplanung']
    date_col = {}
    for col in range(2, ws.max_column + 1):
        dt = ws.cell(row=3, column=col).value
        if dt:
            date_col[dt.date()] = col
    return (ws, date_col)

def old_individual_day_session(old_idx, dt, catalog, ex, cats_counter):
    if not old_idx or not dt or dt not in old_idx[1]:
        return None
    ws_old, date_col = old_idx
    col = date_col[dt]
    ort = ws_old.cell(row=OLD_IND_ORT_ROW, column=col).value
    cats = []
    any_item = False
    for cat_name in CAT_ORDER:
        items = []
        for r in OLD_IND_CAT_ROWS[cat_name]:
            val = ws_old.cell(row=r, column=col).value
            if val:
                m = fuzzy_match(str(val), catalog, ex) if cat_name != 'Technikfokus' else None
                items.append({'ex': str(val), 'match': m})
                any_item = True
                cats_counter[cat_name] = cats_counter.get(cat_name, 0) + 1
        cats.append({'name': cat_name, 'items': items})
    notiz_parts = [str(ws_old.cell(row=r, column=col).value).strip()
                   for r in OLD_IND_NOTIZ_ROWS if ws_old.cell(row=r, column=col).value]
    notiz = ' / '.join(notiz_parts) if notiz_parts else None
    if not (ort or any_item or notiz):
        return None
    return {'name': 'Session 1', 'time': '', 'ort': ort, 'cats': cats, 'notiz': notiz}

# Gruppen-Einheitenplanung: absolute Zeilen (siehe Kopfzeilen-Dump U9).
# Korrektur 27.08.: "Physis Wand" (14,15) und "Physis Boden" (16,17) sind laut
# Floyd getrennte Kategorien (Physis Wand -> Physis Klettern, Physis Boden ->
# Athletik + Mentales) — vorher fälschlich beide in Physis Klettern zusammengefasst,
# wodurch Athletik + Mentales bei Gruppen immer leer war. Motorik gibt es in der
# Gruppenplanung nicht -> immer "-".
# Dieses Raster gilt für das ALTE "Einheitenplanung"-Sheet (Saison 25/26).
GRP_CAT_ROWS_OLD = {
    'Theorie': [8], 'Aufwärmen': [9, 10], 'Motorik': [], 'Technikfokus': [11],
    'Technik/Taktik': [12, 13], 'Physis Klettern': [14, 15], 'Athletik + Mentales': [16, 17],
    'Spezial': [18, 19],
}
GRP_NOTIZ_ROWS_OLD = [21, 22]
GRP_HEADER_ROWS_OLD = {'trainer': 5, 'zeit': 6, 'ort': 7}

# 28.08.: Floyd hat "Einheitenplanung 26_27" für Gruppen neu strukturiert -
# Uhrzeit/Trainer-Zeilen ergänzt und das Kategorienraster an die Einzelpläne
# angeglichen (jetzt inkl. echter Motorik-Zeile, Athletik nur noch 1 Zeile).
# Dadurch verschieben sich ALLE Zeilennummern gegenüber dem alten Sheet.
GRP_CAT_ROWS_NEW = {
    'Theorie': [9], 'Aufwärmen': [10], 'Motorik': [11], 'Technikfokus': [12],
    'Technik/Taktik': [13, 14], 'Physis Klettern': [15, 16], 'Athletik + Mentales': [17],
    'Spezial': [19],
}
GRP_NOTIZ_ROWS_NEW = [18, 20]
GRP_HEADER_ROWS_NEW = {'trainer': 7, 'zeit': 6, 'ort': 8}

# Rückwärtskompatible Aliase (Default = altes Raster, falls irgendwo ohne
# Schema-Auswahl referenziert).
GRP_CAT_ROWS = GRP_CAT_ROWS_OLD
GRP_NOTIZ_ROWS = GRP_NOTIZ_ROWS_OLD

# ------------------------------------------- Saison-Statistik (25/26 + 26/27) -
# Eigene Sheets ("Einheitenplanung" ohne 26_27-Zusatz) mit eigenem Zeilenraster
# (Skill/Physis Wand/Physis Boden statt Technikfokus/Technik-Taktik/Physis Klettern
# etc.) — das ist die abgeschlossene Saison 25/26, getrennt vom laufenden Live-Plan.
# Die Statistik kann zwischen Saison 25/26 (abgeschlossen) und 26/27 (laufend/
# kommend) umgeschaltet werden — Woche/Jahr/Benchmarks bleiben davon unberührt
# und zeigen weiterhin fortlaufend den Live-Stand ohne Saison-Umbruch.
SEASON_LABEL = '25/26'
SEASONS = ['25/26', '26/27']
SEASON_WEEKS = set([(2025, k) for k in range(38, 53)] + [(2026, k) for k in range(1, 38)])
SEASON_WEEKS_2627 = set([(2026, k) for k in range(38, 53)] + [(2027, k) for k in range(1, 38)])

IND_SEASON_CAT_ROWS = {
    'Theorie': [6], 'Aufwärmen': [7, 8], 'Motorik': [], 'Technikfokus': [9],
    'Technik/Taktik': [10, 11, 12], 'Physis Klettern': [13, 14, 15],
    'Athletik + Mentales': [16, 17, 18, 19, 20], 'Spezial': [21, 22, 23],
}
GRP_SEASON_CAT_ROWS = GRP_CAT_ROWS_OLD  # Saison 25/26 liest das alte Sheet/Raster

# Gruppen tracken keine echte Trainingsdauer (kein Dauer-Feld in der Gruppenplanung).
# Floyds Vorgabe: für die Statistik pauschal 2,75h pro echtem Trainingstag annehmen.
# Ein Trainingstag = mindestens eine der Kategorie-Zeilen unterhalb "Trainingsort"
# hat an dem Tag einen Eintrag (Notizen zählen nicht). Klar als Schätzung ausgewiesen,
# nicht als gemessene Dauer.
GROUP_EST_HOURS_PER_DAY = 2.75

# ------------------------------------------------------- Einheitenplanung --

def iso(d):
    if not d:
        return None
    c = d.isocalendar()
    return (c[0], c[1])

def parse_individual(path, athlete_name, catalog, ex, cats_counter):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheetname = None
    for cand in ['Einheitenplanung 26_27']:
        if cand in wb.sheetnames:
            sheetname = cand
            break
    if not sheetname:
        return {}, None
    ws = wb[sheetname]
    maxc = ws.max_column
    row_labels = {r: ws.cell(row=r, column=1).value for r in range(1, 41)}
    weeks = {}   # (year,kw) -> {days:[...], span, source}
    week_stats = {}  # (year,kw) -> accumulate load/hrs/rpe/fit/days
    old_idx = build_old_individual_index(wb)

    c = 2
    while c <= maxc:
        kw_val = ws.cell(row=1, column=c).value
        if not isinstance(kw_val, (int, float)):
            c += 1
            continue
        block_dates = []
        for i in range(7):
            dt = ws.cell(row=3, column=c + i).value
            block_dates.append(dt.date() if dt else None)
        first_dt = next((d for d in block_dates if d), None)
        if not first_dt:
            c += 7
            continue
        yk = iso(first_dt)
        days = []
        for i in range(7):
            col = c + i
            dow = ws.cell(row=2, column=col).value
            dt = block_dates[i]
            sessions = []
            for sess_num, base in [(1, 4), (2, 19)]:
                ort = ws.cell(row=base + 2, column=col).value
                cats = []
                any_item = False
                for cat_name in CAT_ORDER:
                    items = []
                    for off in IND_CAT_ROWS[cat_name]:
                        val = ws.cell(row=base + off, column=col).value
                        if val:
                            m = fuzzy_match(str(val), catalog, ex) if cat_name != 'Technikfokus' else None
                            items.append({'ex': str(val), 'match': m})
                            any_item = True
                            cats_counter[cat_name] = cats_counter.get(cat_name, 0) + 1
                    cats.append({'name': cat_name, 'items': items})
                notiz_parts = [str(ws.cell(row=base + off, column=col).value).strip()
                               for off in IND_NOTIZ_OFFSETS if ws.cell(row=base + off, column=col).value]
                notiz = ' / '.join(notiz_parts) if notiz_parts else None
                if ort or any_item or notiz:
                    sessions.append({'name': f'Session {sess_num}', 'time': '', 'ort': ort, 'cats': cats, 'notiz': notiz})
            if not sessions:
                fallback = old_individual_day_session(old_idx, dt, catalog, ex, cats_counter)
                if fallback:
                    sessions.append(fallback)
            mot = ws.cell(row=35, column=col).value
            eb = ws.cell(row=36, column=col).value
            ee = ws.cell(row=37, column=col).value
            d1 = ws.cell(row=38, column=col).value
            d2 = ws.cell(row=39, column=col).value
            doku = None
            if any(v is not None for v in (mot, eb, ee, d1, d2)):
                doku = {'mot': mot, 'eb': eb, 'ee': ee, 'd1': d1, 'd2': d2}
                hrs = (float(d1) if isinstance(d1, (int, float)) else 0) + (float(d2) if isinstance(d2, (int, float)) else 0)
                load = hrs * ee if isinstance(ee, (int, float)) else None
                stat = week_stats.setdefault(yk, {'load': 0, 'hrs': 0, 'rpe_sum': 0, 'rpe_n': 0, 'fit_sum': 0, 'fit_n': 0, 'days': 0})
                stat['hrs'] += hrs
                if load is not None:
                    stat['load'] += load
                if isinstance(ee, (int, float)):
                    stat['rpe_sum'] += ee
                    stat['rpe_n'] += 1
                if isinstance(mot, (int, float)):
                    stat['fit_sum'] += mot
                    stat['fit_n'] += 1
                stat['days'] += 1
            days.append({
                'dow': dow, 'dom': dt.strftime('%d.%m.') if dt else '?', 'date': dt.isoformat() if dt else None,
                'ort': next((s['ort'] for s in sessions if s.get('ort')), None),
                'sessions': sessions, 'doku': doku, 'notiz': None,
            })
        span = f"{block_dates[0].strftime('%d.%m.')} – {block_dates[-1].strftime('%d.%m.%Y')}" if block_dates[0] and block_dates[-1] else ''
        weeks[f'{yk[0]}-{yk[1]}'] = {'days': days, 'span': span, 'year': yk[0], 'kw': yk[1]}
        c += 7

    weeks_out = []
    for yk, st in sorted(week_stats.items()):
        weeks_out.append({
            'year': yk[0], 'kw': yk[1],
            'load': round(st['load'], 1), 'hrs': round(st['hrs'], 2),
            'rpe': round(st['rpe_sum'] / st['rpe_n'], 2) if st['rpe_n'] else None,
            'fit': round(st['fit_sum'] / st['fit_n'], 2) if st['fit_n'] else None,
            'days': st['days'],
        })
    return weeks, weeks_out

# Trainingsort-Zellfarbe -> Sessionsart, 1:1 wie von Floyd vorgegeben (dieselben
# Farbcodes wie im Gruppen-Jahresplan "Fokus": Grün=Lead, Blau=Bouldern,
# Gelb=Speed). Kein Erfinden: fehlt die Farbe, bleibt ortType None.
ORT_COLOR_MAP = {'#92D050': 'Lead', '#00B0F0': 'Bouldern', '#FFFF00': 'Speed'}

def parse_group(path, sheet_candidates, catalog, ex, cats_counter):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheetname = None
    for cand in sheet_candidates:
        for sn in wb.sheetnames:
            if sn.strip() == cand:
                sheetname = sn
                break
        if sheetname:
            break
    if not sheetname:
        return {}, sheetname
    ws = wb[sheetname]
    maxc = ws.max_column
    # 28.08.: Floyd hat "Einheitenplanung 26_27" umgebaut (Uhrzeit/Trainer neu,
    # Kategorienraster verschoben) - Zeilenraster hängt daher vom Sheet ab.
    is_new = sheetname.strip() == 'Einheitenplanung 26_27'
    cat_rows = GRP_CAT_ROWS_NEW if is_new else GRP_CAT_ROWS_OLD
    notiz_rows = GRP_NOTIZ_ROWS_NEW if is_new else GRP_NOTIZ_ROWS_OLD
    hdr = GRP_HEADER_ROWS_NEW if is_new else GRP_HEADER_ROWS_OLD
    weeks = {}
    col = 2
    while col <= maxc:
        dt_cell = ws.cell(row=3, column=col).value
        if not dt_cell:
            col += 1
            continue
        dt = dt_cell.date() if hasattr(dt_cell, 'date') else None
        if not dt:
            col += 1
            continue
        yk = iso(dt)
        key = f'{yk[0]}-{yk[1]}'
        wk_entry = weeks.setdefault(key, {'days': [], 'year': yk[0], 'kw': yk[1]})
        dow = ws.cell(row=2, column=col).value
        trainer = ws.cell(row=hdr['trainer'], column=col).value
        zeit = ws.cell(row=hdr['zeit'], column=col).value
        ort_cell = ws.cell(row=hdr['ort'], column=col)
        ort = ort_cell.value
        ort_color = get_fill_hex(ort_cell)
        ort_type = ORT_COLOR_MAP.get(ort_color) if ort_color else None
        cats = []
        any_item = False
        for cat_name in CAT_ORDER:
            items = []
            for r in cat_rows[cat_name]:
                val = ws.cell(row=r, column=col).value
                if val:
                    m = fuzzy_match(str(val), catalog, ex) if cat_name != 'Technikfokus' else None
                    items.append({'ex': str(val), 'match': m})
                    any_item = True
                    cats_counter[cat_name] = cats_counter.get(cat_name, 0) + 1
            cats.append({'name': cat_name, 'items': items})
        notiz_parts = [str(ws.cell(row=r, column=col).value).strip() for r in notiz_rows if ws.cell(row=r, column=col).value]
        notiz = ' / '.join(notiz_parts) if notiz_parts else None
        sessions = []
        if ort or trainer or any_item or notiz:
            sessions.append({'name': 'Einheit', 'time': str(zeit) if zeit else '', 'trainer': trainer, 'ort': ort, 'cats': cats, 'notiz': notiz})
        wk_entry['days'].append({
            'dow': dow, 'dom': dt.strftime('%d.%m.'), 'date': dt.isoformat(),
            'ort': ort, 'ortType': ort_type, 'trainer': trainer, 'zeit': str(zeit) if zeit else None,
            'sessions': sessions, 'doku': None, 'notiz': None,
        })
        col += 1
    for key, wk_entry in weeks.items():
        ds = [d['date'] for d in wk_entry['days'] if d['date']]
        if ds:
            a, b = min(ds), max(ds)
            wk_entry['span'] = f"{date.fromisoformat(a).strftime('%d.%m.')} – {date.fromisoformat(b).strftime('%d.%m.%Y')}"
        else:
            wk_entry['span'] = ''
    return weeks, sheetname

def _find_season_sheet(wb, base_name):
    for cand in [base_name, base_name + ' ', base_name.strip()]:
        if cand in wb.sheetnames:
            return cand
    for sn in wb.sheetnames:
        if sn.strip().lower() == base_name.lower():
            return sn
    return None

def parse_season_individual(path, season_weeks=None):
    """Saison 25/26 aus dem alten (nicht-26_27) Einheitenplanung-Sheet:
    eigenes Zeilenraster (Skill/Physis Wand/Physis Boden), zählt nur echte
    Zelleneinträge pro Kategorie, plus Load/RPE/Dauer/Fitness pro Woche
    (nur dort wo die Zellen wirklich gefüllt sind — kein Erfinden)."""
    season_weeks = season_weeks if season_weeks is not None else SEASON_WEEKS
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    sheetname = _find_season_sheet(wb, 'Einheitenplanung')
    if not sheetname:
        return None
    ws = wb[sheetname]
    maxc = ws.max_column
    cat_counts = {c: 0 for c in CAT_ORDER}
    week_stats = {}
    col = 2
    while col <= maxc:
        kw_val = ws.cell(row=1, column=col).value
        if not isinstance(kw_val, (int, float)):
            col += 1
            continue
        kw = int(kw_val)
        stat = {'load': 0.0, 'hrs': 0.0, 'rpe_sum': 0.0, 'rpe_n': 0, 'fit_sum': 0.0, 'fit_n': 0, 'days': 0}
        week_year = None
        any_real_day = False
        for i in range(7):
            c = col + i
            dt_cell = ws.cell(row=3, column=c).value
            dt = dt_cell.date() if hasattr(dt_cell, 'date') else None
            if not dt:
                continue
            yk = iso(dt)
            if yk not in season_weeks:
                continue
            week_year = yk[0]
            notiz = ws.cell(row=27, column=c).value
            is_example = bool(notiz) and 'beispiel' in str(notiz).lower()
            if is_example:
                continue
            any_real_day = True
            for cat_name, rows in IND_SEASON_CAT_ROWS.items():
                for r in rows:
                    if ws.cell(row=r, column=c).value:
                        cat_counts[cat_name] += 1
            rpe = ws.cell(row=25, column=c).value
            dauer = ws.cell(row=26, column=c).value
            fit = ws.cell(row=28, column=c).value
            load = ws.cell(row=32, column=c).value
            done = ws.cell(row=24, column=c).value
            if isinstance(dauer, (int, float)):
                stat['hrs'] += dauer
            if isinstance(load, (int, float)):
                stat['load'] += load
            elif isinstance(rpe, (int, float)) and isinstance(dauer, (int, float)):
                stat['load'] += rpe * dauer
            if isinstance(rpe, (int, float)):
                stat['rpe_sum'] += rpe
                stat['rpe_n'] += 1
            if isinstance(fit, (int, float)):
                stat['fit_sum'] += fit
                stat['fit_n'] += 1
            if done or isinstance(rpe, (int, float)) or isinstance(dauer, (int, float)):
                stat['days'] += 1
        if any_real_day and week_year is not None:
            week_stats[(week_year, kw)] = stat
        col += 7

    weeks_out = []
    for (y, kw), st in sorted(week_stats.items()):
        weeks_out.append({
            'year': y, 'kw': kw,
            'load': round(st['load'], 1) if st['load'] else (0 if st['days'] else None),
            'hrs': round(st['hrs'], 2),
            'rpe': round(st['rpe_sum'] / st['rpe_n'], 2) if st['rpe_n'] else None,
            'fit': round(st['fit_sum'] / st['fit_n'], 2) if st['fit_n'] else None,
            'days': st['days'],
        })
    return {
        'catCounts': [{'name': c, 'count': cat_counts[c]} for c in CAT_ORDER],
        'weeks': weeks_out,
        'sheet': sheetname,
    }

def build_group_season_stats_from_weeks(weeks, season_weeks, sheet_label):
    """Saison 26/27 für Gruppen: direkt aus den bereits per parse_group()
    geparsten Einheitenplanung-26_27-Daten (neues Zeilenraster inkl. Motorik)
    aggregiert, statt nochmal mit dem alten Sheet/Raster zu parsen."""
    cat_counts = {c: 0 for c in CAT_ORDER}
    training_days = 0
    for wk in weeks.values():
        if (wk.get('year'), wk.get('kw')) not in season_weeks:
            continue
        for day in wk.get('days', []):
            day_has_entry = False
            for sess in day.get('sessions', []):
                for cat in sess.get('cats', []):
                    n = len(cat.get('items') or [])
                    if n:
                        cat_counts[cat['name']] = cat_counts.get(cat['name'], 0) + n
                        day_has_entry = True
            if day_has_entry:
                training_days += 1
    if training_days == 0 and all(v == 0 for v in cat_counts.values()):
        return None
    return {
        'catCounts': [{'name': c, 'count': cat_counts.get(c, 0)} for c in CAT_ORDER],
        'weeks': [],
        'sheet': sheet_label,
        'trainingDays': training_days,
        'estHoursPerDay': GROUP_EST_HOURS_PER_DAY,
        'estHours': round(training_days * GROUP_EST_HOURS_PER_DAY, 1),
    }

def parse_season_group(path, season_weeks=None):
    """Saison für Gruppen: dasselbe Sheet wie der Live-Plan (Gruppen haben
    kein separates 26_27-Sheet), nur auf die jeweilige Saison gefiltert.
    Kein Load möglich (keine RPE/Dauer-Zeilen in der Gruppenplanung)."""
    season_weeks = season_weeks if season_weeks is not None else SEASON_WEEKS
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    sheetname = _find_season_sheet(wb, 'Einheitenplanung')
    if not sheetname:
        return None
    ws = wb[sheetname]
    maxc = ws.max_column
    cat_counts = {c: 0 for c in CAT_ORDER}
    training_days = 0
    all_cat_rows = sorted(set(r for rows in GRP_SEASON_CAT_ROWS.values() for r in rows))
    for col in range(2, maxc + 1):
        dt_cell = ws.cell(row=3, column=col).value
        dt = dt_cell.date() if hasattr(dt_cell, 'date') else None
        if not dt:
            continue
        if iso(dt) not in season_weeks:
            continue
        day_has_entry = False
        for cat_name, rows in GRP_SEASON_CAT_ROWS.items():
            for r in rows:
                if ws.cell(row=r, column=col).value:
                    cat_counts[cat_name] += 1
        for r in all_cat_rows:
            if ws.cell(row=r, column=col).value:
                day_has_entry = True
                break
        if day_has_entry:
            training_days += 1
    return {
        'catCounts': [{'name': c, 'count': cat_counts[c]} for c in CAT_ORDER],
        'weeks': [],
        'sheet': sheetname,
        'trainingDays': training_days,
        'estHoursPerDay': GROUP_EST_HOURS_PER_DAY,
        'estHours': round(training_days * GROUP_EST_HOURS_PER_DAY, 1),
    }

# -------------------------------------------------------------- Benchmarks --

BENCH_STRUCTURE = [
    ('Technik+Taktik', 'mitlaufend', [
        'Rotpunkt K1', 'Onsight K1', 'Rotpunkt Fels', 'Onsight Fels',
        'Kilterboard Max', 'Kilterboard Flash', 'Kilterboard Base', 'Steinblock DB Max',
    ]),
    ('Motorik', 'mitlaufend', ['Anzahl Bälle Jonglieren', 'Alternate Wall Toss']),
    ('Physisches Klettertraining', 'mitlaufend', ['Anzahl Aufbauboulder Sessions', 'Doppellänge Max']),
    ('Athletik', 'regelmäßig', [
        '90° Block Einarmig', 'Half Crimp langer Arm', 'PinchPower', 'Einarmer Ja/Nein',
        'Klimmzug Kraftausdauer', 'Spagat Seite Abstand Wand', 'Jump and Reach Test', 'Handstand frei in Sekunden',
    ]),
]

def load_bench_source():
    """Liest die bestehende Benchmarks.xlsx (nur Adrian real gepflegt)."""
    wb = openpyxl.load_workbook(BASE + 'Zusatzinfos/Benchmarks.xlsx', data_only=True)
    ws = wb['Tabelle1']
    vals = {}
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=3).value
        v = ws.cell(row=r, column=4).value
        if item:
            key = norm(re.sub(r'^[-\d.\s]+', '', str(item)))
            vals[key] = v
    return vals

def build_bench_for(athlete_name, src_vals, aufbau_count=None):
    cats_out = []
    any_val = False
    for cat_name, rhythm, items in BENCH_STRUCTURE:
        its = []
        for it in items:
            auto = 'Sessions' in it
            if auto:
                v = aufbau_count
            else:
                v = src_vals.get(norm(it)) if src_vals else None
            if v is not None:
                any_val = True
            its.append({'k': it, 'v': v if v is not None else None, 'p': None, 't': 0, 'auto': auto})
        cats_out.append({'name': cat_name, 'rhythm': rhythm, 'items': its})
    if not any_val:
        return None
    return {'stand': '27.08.2026', 'cats': cats_out}

def count_aufbau_sessions(weeks):
    """Zählt reale Vorkommen von 'Aufbau Boulder'-Übungen (gematcht über die
    Übungssammlung) über alle Wochen einer Athletin/eines Athleten."""
    n = 0
    for wk in weeks.values():
        for day in wk['days']:
            for sess in day['sessions']:
                for cat in sess.get('cats', []):
                    for it in cat['items']:
                        m = it.get('match')
                        if m and str(m.get('name', '')).startswith('Aufbau Boulder'):
                            n += 1
    return n

# ---------------------------------------------------------------- Kalender --

AGE_CLASSES = ['U9', 'U11', 'U13', 'U15', 'U17', 'U19']

def load_gk():
    wb = openpyxl.load_workbook(BASE + 'Zusatzinfos/Gesamtkalender 2025_2026 – aktuell.xlsx', data_only=True)
    ws = wb['MOAP']
    rows = []
    for r in range(3, 17):
        label = ws.cell(row=r, column=1).value
        if label:
            rows.append((r, str(label)))
    weeks_range = {}
    for c in range(4, ws.max_column + 1):
        rng = ws.cell(row=2, column=c).value
        if not rng:
            continue
        weeks_range[c] = str(rng)

    def parse_start_date(rng, ref_year):
        m = re.findall(r'(\d+)\.(\d+)?\.?', rng.split('-')[0])
        try:
            day = int(re.search(r'^(\d+)\.', rng).group(1))
            monthmatch = re.search(r'\.(\d+)\.', rng.split('-')[0])
            month = int(monthmatch.group(1)) if monthmatch else None
            return day, month
        except Exception:
            return None, None

    gk = {a: {} for a in AGE_CLASSES}
    cur_year = 2025
    last_month = 8
    for c, rng in weeks_range.items():
        day, month = parse_start_date(rng, cur_year)
        if month is not None:
            if month < last_month - 6:
                cur_year += 1
            last_month = month
        try:
            dt = date(cur_year, month, day) if month and day else None
        except Exception:
            dt = None
        yk = iso(dt) if dt else None
        for (r, label) in rows:
            val = ws.cell(row=r, column=c).value
            if not val:
                continue
            m = re.findall(r'U(9|11|13|15|17|19)', label)
            classes = [f'U{x}' for x in m] if m else AGE_CLASSES[:]
            for cl in classes:
                if yk:
                    key = f'{yk[0]}-{yk[1]}'
                    entry = gk[cl].setdefault(key, {'d': rng, 'items': []})
                    entry['items'].append({'c': label, 't': str(val)})
    return gk

# ------------------------------------------------------------ Jahresplanung --

# Bekannte kurze Phasen-Schlagworte -> normalisierte Kategorie (nur diese werden
# im Phasen-Ribbon eingefärbt; lange Freitext-Beschreibungen bleiben unfarbig
# und werden nur als Text angezeigt, statt sie zu erraten / zu erfinden).
def normalize_phase(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    low = s.lower()
    if low.startswith('max/sk') or low.startswith('max / sk'):
        return 'Max/SK'
    if low.startswith('fels + basis') or low.startswith('fels+basis'):
        return 'Fels + Basis'
    if low.startswith('frei/pause') or low.startswith('frei / pause') or low == 'pause':
        return 'Pause'
    if low.startswith('basis'):
        return 'Basis'
    if low.startswith('aufbau'):
        return 'Aufbau'
    if low.startswith('skill'):
        return 'Skill'
    if low.startswith('fels'):
        return 'Fels'
    if low.startswith('max'):
        return 'Max'
    if low == 'sk':
        return 'SK'
    if low == 'ka':
        return 'KA'
    if low == 'wk':
        return 'WK'
    return None

def get_fill_hex(cell):
    """Echte Zellfarbe (ARGB) als '#rrggbb' oder None (keine/transparente Füllung)."""
    try:
        fg = cell.fill.fgColor
        rgb = fg.rgb if fg else None
    except Exception:
        rgb = None
    if not isinstance(rgb, str) or len(rgb) != 8:
        return None
    if rgb[:2] == '00':
        return None
    return '#' + rgb[2:]

def parse_jahresplanung(path):
    """Liest den echten Jahresplan (Phase/Trainingsumfang/WK-Termine Boulder+Lead/
    Fels/Schulferien/Sonstiges) aus dem Reiter 'Jahresplanung' einer Athlet:innen-
    Excel. Zeilen-Offsets variieren leicht pro Datei, daher werden sie über die
    Beschriftung in Spalte A gefunden statt über feste Zeilennummern."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    if 'Jahresplanung' not in wb.sheetnames:
        return None
    ws = wb['Jahresplanung']

    kw_row = None
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith('kalenderwoche'):
            kw_row = r
            break
    if kw_row is None:
        return None
    phase_row = kw_row + 1
    umfang_row = kw_row + 2

    wk_row = None
    for r in range(umfang_row + 1, umfang_row + 8):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith('wk termine'):
            wk_row = r
            break
    if wk_row is None:
        return None
    bouldern_row = wk_row
    lead_row = wk_row + 1
    fels_row = lead_row + 1
    ferien_row = fels_row + 1
    sonstiges_row = ferien_row + 1
    note_rows = list(range(umfang_row + 1, wk_row))

    title = str(ws.cell(row=1, column=3).value or '')
    m = re.search(r'(20\d{2})', title)
    start_year = int(m.group(1)) if m else date.today().year

    def cellstr(row, col):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        s = str(v).strip()
        return s if s and s.lower() != 'leer' else None

    # Farbe->Label-Zuordnung über ALLE Spalten vorab lernen (auch die reinen
    # Monats-Trennspalten ohne numerische KW) - Floyd trägt die Phasen-
    # Bezeichnung manchmal genau auf so einer Trennspalte ein, während die
    # Farbe sich schon auf die folgenden echten KW-Spalten fortsetzt.
    color_label_map = {}
    for c in range(3, ws.max_column + 1):
        praw = cellstr(phase_row, c)
        pnorm = normalize_phase(praw)
        pcolor = get_fill_hex(ws.cell(row=phase_row, column=c))
        if pnorm and pcolor:
            color_label_map[pcolor] = pnorm

    weeks = []
    year = start_year
    prev_kw = None
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=kw_row, column=c).value
        if not isinstance(v, (int, float)):
            continue
        kw = int(v)
        if prev_kw is not None and kw < prev_kw - 5:
            year += 1
        prev_kw = kw
        phase_raw = cellstr(phase_row, c)
        phase_color = get_fill_hex(ws.cell(row=phase_row, column=c))
        phase_norm = normalize_phase(phase_raw)
        if phase_norm and phase_color:
            color_label_map[phase_color] = phase_norm
        # Farbe nur für WIRKLICH leere Zellen (keine eigene Beschriftung)
        # von einer anderen Woche gleicher Farbe übernehmen. Steht in der
        # Zelle bereits eigener (auch unbekannter/freitextiger) Text, wird
        # der nie durch eine geratene Kurzphase ersetzt.
        if phase_raw:
            resolved_phase = phase_norm
        else:
            resolved_phase = color_label_map.get(phase_color) if phase_color else None
        umfang = cellstr(umfang_row, c)
        notes = []
        for nr in note_rows:
            label = str(ws.cell(row=nr, column=1).value or '').strip().rstrip(':') or 'Notiz'
            txt = cellstr(nr, c)
            if txt:
                notes.append({'label': label, 'text': txt})
        weeks.append({
            'year': year, 'kw': kw,
            'phase': resolved_phase,
            'phaseRaw': phase_raw,
            'phaseColor': phase_color,
            'reduziert': bool(umfang and umfang.lower().startswith('red')),
            'boulder': cellstr(bouldern_row, c),
            'lead': cellstr(lead_row, c),
            'fels': cellstr(fels_row, c),
            'ferien': cellstr(ferien_row, c),
            'sonstiges': cellstr(sonstiges_row, c),
            'notes': notes,
        })
    return weeks or None

def parse_group_jahresplanung(path):
    """Liest den echten Fokus-Verlauf (Lead/Bouldern/Speed, farbcodiert) aus dem
    Reiter 'Jahresplan <Gruppe>.xlsx'. Analog zu parse_jahresplanung, aber mit
    dem einfacheren Gruppen-Zeilenraster (nur 'Fokus'-Zeile statt 'Phase')."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return None
    ws = wb[wb.sheetnames[0]]

    kw_row = None
    for r in range(1, 10):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith('kalenderwoche'):
            kw_row = r
            break
    if kw_row is None:
        return None

    # Es kann mehrere "Fokus..."-Zeilen geben (z.B. "Fokus Stand 08.2025" und
    # "Fokus Stand 01.2026") - die mit den meisten echten Einträgen gewinnt.
    fokus_candidates = []
    for r in range(kw_row + 1, kw_row + 5):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith('fokus'):
            fokus_candidates.append(r)
    if not fokus_candidates:
        return None

    def cellstr(row, col):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        s = str(v).strip()
        return s if s and s.lower() != 'leer' else None

    def count_filled(r):
        return sum(1 for c in range(3, ws.max_column + 1) if cellstr(r, c))
    phase_row = max(fokus_candidates, key=count_filled)

    start_year = 2025
    for c in range(1, 10):
        v = ws.cell(row=1, column=c).value
        if v:
            m = re.search(r'(20\d{2})', str(v))
            if m:
                start_year = int(m.group(1))
                break

    # Farbe->Label über ALLE Spalten vorab lernen (auch Monats-Trennspalten
    # ohne numerische KW, auf denen das Label manchmal allein sitzt).
    color_label_map = {}
    for c in range(3, ws.max_column + 1):
        praw = cellstr(phase_row, c)
        pcolor = get_fill_hex(ws.cell(row=phase_row, column=c))
        if praw and pcolor:
            color_label_map[pcolor] = praw

    weeks = []
    year = start_year
    prev_kw = None
    for c in range(3, ws.max_column + 1):
        v = ws.cell(row=kw_row, column=c).value
        if not isinstance(v, (int, float)):
            continue
        kw = int(v)
        if prev_kw is not None and kw < prev_kw - 5:
            year += 1
        prev_kw = kw
        phase_raw = cellstr(phase_row, c)
        phase_color = get_fill_hex(ws.cell(row=phase_row, column=c))
        resolved_phase = phase_raw or (color_label_map.get(phase_color) if phase_color else None)
        weeks.append({
            'year': year, 'kw': kw,
            'phase': resolved_phase,
            'phaseRaw': phase_raw,
            'phaseColor': phase_color,
        })
    return weeks or None

# ----------------------------------------------------------------- Coaches --

COACHES_EXTRA = [
    {'n': 'Matteo', 'r': 'Coach', 'note': 'Athletik → U13/U15'},
    {'n': 'Naima', 'r': 'Coach', 'note': 'Springerin'},
]
HEADCOACHES = {'Floyd Simen', 'Mark Amann'}

def build_coaches(groups):
    seen = {}
    for g in groups:
        names = re.split(r'\s*\+\s*', g['coaches']) if g['coaches'] else []
        for n in names:
            n = n.strip()
            if not n:
                continue
            seen[n] = 'Headcoach' if n in HEADCOACHES else 'Coach'
    coaches = [{'n': n, 'r': r} for n, r in seen.items()]
    for extra in COACHES_EXTRA:
        if extra['n'] not in [c['n'] for c in coaches]:
            coaches.append({'n': extra['n'], 'r': 'Coach'})
    return coaches

# =============================================================== MAIN =====

def main():
    ex = load_ex()
    catalog = build_catalog(ex)
    groups_kader = load_kader()
    coaches = build_coaches(groups_kader)
    bench_src = load_bench_source()

    INDIVIDUAL_FILES = {
        'Adrian Kathan': 'Adrian Kathan 26_27.xlsx',
        'Mariella Vierhauser': 'Mariella Vierhauser 26_27.xlsx',
        'Raphael Hubmann': 'Raphael Hubmann 26_27.xlsx',
        'Jakob Burtscher': 'Jakob Burtscher 26_27.xlsx',
        'Linus Pfleger': 'Linus Pfleger 26_27.xlsx',
        'Sophie Bickel': 'Sophie Bickel 26_27.xlsx',
        'Matthäus Kathan': 'Matthäus Kathan 26_27.xlsx',
        'Levi Strolz': 'Levi Strolz 26_27.xlsx',
    }
    # Individuen liegen jetzt im Unterordner "Einzelpläne/" (Floyd hat den
    # Testspace-Ordner neu strukturiert); resolve_file fällt bei Bedarf auf
    # den alten Wurzelpfad zurück und toleriert OneDrive-Konfliktkopien
    # (z.B. "Name 2.xlsx").
    INDIVIDUAL_PATHS = {
        name: resolve_file(BASE + 'Einzelpläne/' + fname, BASE + fname)
        for name, fname in INDIVIDUAL_FILES.items()
    }
    GROUP_FILES = {
        'U9': ('Gruppenpläne/U9/Einheitenplanung U9.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U11I': ('Gruppenpläne/U11 I/Einheitenplanung U11 I.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U11II': ('Gruppenpläne/U11 II/Einheitenplanung U11 II.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U13 I': ('Gruppenpläne/U13 I/Einheitenplanung U13 I.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U13II': ('Gruppenpläne/U13 II/Einheitenplanung U13 II.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U15I': ('Gruppenpläne/U15 I/EINHEITENPLANUNG U15 I.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
        'U15 II': ('Gruppenpläne/U15 II/EINHEITENPLANUNG U15 II.xlsx', ['Einheitenplanung 26_27', 'Einheitenplanung']),
    }
    GROUP_JAHRESPLAN_FILES = {
        'U9': 'Gruppenpläne/U9/Jahresplan U9.xlsx',
        'U11I': 'Gruppenpläne/U11 I/Jahresplan U11 I.xlsx',
        'U11II': 'Gruppenpläne/U11 II/Jahresplan U11 II.xlsx',
        'U13 I': 'Gruppenpläne/U13 I/Jahresplan U13 I.xlsx',
        'U13II': 'Gruppenpläne/U13 II/Jahresplan U13 II.xlsx',
        'U15I': 'Gruppenpläne/U15 I/Jahresplanung U15.xlsx',
        'U15 II': 'Gruppenpläne/U15 II/Jahresplanung U15.xlsx',
    }

    PLANS = {}
    WEEKS_BY_ATHLETE = {}
    CATS_BY_ATHLETE = {}
    BENCH = {}
    SOURCE_INFO = {}
    JAHRESPLAN = {}
    JAHRESPLAN_SOURCE = {}
    JAHRESPLAN_GROUP = {}
    JAHRESPLAN_GROUP_SOURCE = {}
    SEASON_STATS = {'25/26': {}, '26/27': {}}

    for name, fname in INDIVIDUAL_FILES.items():
        fpath = INDIVIDUAL_PATHS[name]
        cats_counter = {}
        weeks, weeks_out = parse_individual(fpath, name, catalog, ex, cats_counter)
        PLANS['a:' + name] = weeks
        WEEKS_BY_ATHLETE[name] = weeks_out
        CATS_BY_ATHLETE['a:' + name] = cats_counter
        SOURCE_INFO['a:' + name] = 'Einheitenplanung 26_27 · ' + os.path.basename(fpath)
        aufbau_n = count_aufbau_sessions(weeks)
        b = build_bench_for(name, bench_src if name == 'Adrian Kathan' else None, aufbau_count=(aufbau_n or None))
        if b:
            BENCH[name] = b
        jp = parse_jahresplanung(fpath)
        if jp:
            JAHRESPLAN[name] = jp
            JAHRESPLAN_SOURCE[name] = 'Jahresplanung · ' + os.path.basename(fpath)
        ss = parse_season_individual(fpath)
        if ss:
            SEASON_STATS['25/26']['a:' + name] = ss
        # Saison 26/27: direkt aus den bereits geparsten Einheitenplanung-26_27-Daten
        # (parse_season_individual liest das alte alte Schema und ist für 26/27 nicht gültig)
        if cats_counter or weeks_out:
            SEASON_STATS['26/27']['a:' + name] = {
                'catCounts': [{'name': c, 'count': cats_counter.get(c, 0)} for c in CAT_ORDER],
                'weeks': weeks_out,
                'sheet': 'Einheitenplanung 26_27',
            }

    for gid, (relpath, sheet_cands) in GROUP_FILES.items():
        cats_counter = {}
        weeks, sheetname = parse_group(BASE + relpath, sheet_cands, catalog, ex, cats_counter)
        PLANS['g:' + gid] = weeks
        CATS_BY_ATHLETE['g:' + gid] = cats_counter
        SOURCE_INFO['g:' + gid] = f'{sheetname or "?"} · {relpath.split("/")[-1]}'
        ss = parse_season_group(BASE + relpath, SEASON_WEEKS)
        if ss:
            SEASON_STATS['25/26']['g:' + gid] = ss
        ss27 = build_group_season_stats_from_weeks(weeks, SEASON_WEEKS_2627, sheetname or 'Einheitenplanung 26_27')
        if ss27:
            SEASON_STATS['26/27']['g:' + gid] = ss27
        jp_path = GROUP_JAHRESPLAN_FILES.get(gid)
        if jp_path:
            jpg = parse_group_jahresplanung(BASE + jp_path)
            if jpg:
                JAHRESPLAN_GROUP[gid] = jpg
                JAHRESPLAN_GROUP_SOURCE[gid] = 'Jahresplan · ' + jp_path.split('/')[-1]

    # Missing sources -> explicit empty markers (kein Erfinden)
    for missing in ['Levi Strolz']:
        if ('a:' + missing) not in PLANS:
            PLANS['a:' + missing] = {}
            SOURCE_INFO['a:' + missing] = 'Keine Excel-Datei im Testspace-Ordner hinterlegt'
    if 'g:SPOGY Gruppe' not in PLANS:
        PLANS['g:SPOGY Gruppe'] = {}
        SOURCE_INFO['g:SPOGY Gruppe'] = 'Keine Gruppen-Excel für SPOGY im Testspace-Ordner hinterlegt (Athlet:innen haben Einzelpläne)'

    gk = load_gk()

    data = {
        'GENERATED': '2026-08-27',
        'GROUPS': groups_kader,
        'COACHES': coaches,
        'EX': ex,
        'PLANS': PLANS,
        'CATS_BY_VIEW': {k: [{'name': c, 'count': v.get(c, 0)} for c in CAT_ORDER] for k, v in CATS_BY_ATHLETE.items()},
        'WEEKS_BY_ATHLETE': WEEKS_BY_ATHLETE,
        'BENCH': BENCH,
        'GK': gk,
        'SOURCE_INFO': SOURCE_INFO,
        'JAHRESPLAN': JAHRESPLAN,
        'JAHRESPLAN_SOURCE': JAHRESPLAN_SOURCE,
        'JAHRESPLAN_GROUP': JAHRESPLAN_GROUP,
        'JAHRESPLAN_GROUP_SOURCE': JAHRESPLAN_GROUP_SOURCE,
        'SEASON_STATS': SEASON_STATS,
        'SEASON_LABEL': SEASON_LABEL,
        'SEASONS': SEASONS,
    }

    js = 'window.TPDATA = ' + json.dumps(data, ensure_ascii=False, default=str) + ';\n'
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(js)
    print('wrote', OUT, len(js), 'bytes')
    print('athletes with plans:', {k: len(v) for k, v in PLANS.items() if k.startswith('a:')})
    print('groups with plans:', {k: len(v) for k, v in PLANS.items() if k.startswith('g:')})
    print('bench:', list(BENCH.keys()))

if __name__ == '__main__':
    main()
